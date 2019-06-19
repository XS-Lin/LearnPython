# 機能:
#   ネットワーク通信により、
#     Excel読み取りのリクエスト受信し、Excelの内容を戻します。
#     Excel書き込みのリクエスト受信し、Excelの内容を出力します。
#   なお、Officeのインストールする必要はありません。
# 使用例:
#   WebApplication等と組み合わせ、ServerサイドでExcelを処理します。
#   起動コマンド(python3はPython3.7のシンボリックリンク)
#     python3 excel_tool_by_openpyxl.py
#   テストマンド(正常の場合は「Waiting」)
#     例:base
#       curl http://ip:PORT/status
#     例:ブラウザ
#       http://ip:PORT/statusを開く
#   処理コマンド(HTTPリクエスト送信)
#     例:bash
#       curl -H 'Content-Type:application/json; charset=utf-8' -d '{"key":"70949d2b88c04e869b074597700dd22d","fileName":"test.xml","process_method":"fun1","param":"xxxxxxx"}' http://ip:PORT/read
#     例:bash
#       curl -H 'Content-Type:application/json; charset=utf-8' -d "{"key":"70949d2b88c04e869b074597700dd22d","fileName":"test.xml","process_method":"fun2","param":"xxxxxxx"}" http://ip:PORT/write
#   停止コマンド
#     ctrl+c
# 検証環境
#   CentOs7.5
#   Python3.7.3
#   Windows10
#   Excel2019
# 注意:
#   1.大きいサイズのExcel(例えば100M以上)は想定していません。
#   2.入力ExcelはExcel2010以上が必要になります。(xlsx)
#   3.出力のExcelはExcel2010以上で表示できます。
#   4.使用範囲は内部ネットワークと想定しています。
#   5.デフォルト処理以外の場合は拡張が必要になります。
#   6.動作確認しましたが、大きいサイズやインジェクションのテストは実施していません。
import os
import openpyxl
import urllib.parse
import html
import json
import platform

from copy import copy
from http.server import BaseHTTPRequestHandler
from http.server import ThreadingHTTPServer
from http import HTTPStatus
# -----------------------------------------------------------------------------
# アクセスキー(ヘッダにキーが未設定のtest以外のリクエストは無視する)
ACCESS_KEY = "70949d2b88c04e869b074597700dd22d"
# コマンド受信ポート
PORT = 10001
# ファイル読み取り時のヘーズフォルダ(外層はアクセスしない)
BASE_FOLDER = "/var/tmp"
# 中間ファイル保存用フォルダ
WORK_FOLDER = "/"
# 入力ファイルのフォルダ(WebAppのアップロードフォルダ)
INPUT_FOLDER = "UPLOAD"
# 出力ファイルのフォルダ(WebAppのダウンロード用フォルダ)
OUTPUT_FOLDER = "DOWNLOAD"
# 最大リクエストサイズ1M
MAX_BYTES = 1 * 1024 * 1024
# -----------------------------------------------------------------------------
# 各処理関数の戻り値はstringと想定(文字列またはシリアライズドオブジェクト)
class Action:
    def __init__(self):
        return
    def test_excel_function(self):
        excelFileName = 'for_test'
        excelFileFullName = os.path.join(BASE_FOLDER,excelFileName + '.xlsx')
        try:
            if (os.path.isfile(excelFileFullName)):
                wb = openpyxl.load_workbook(excelFileFullName)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
            ws['A1'].value = 1
            ws.cell(row=2, column=1).value = "123"
            wb.save(excelFileFullName)
        except Exception as e:
            return e.message
        return 'Test success'
    def test_read_excel_function(self,param):
        '''Read data to json from excel identified by param like {"book"="read_excel.xlsx","sheet"="test","range"="A2:C3","header"="a,b,c"}'''
        parameters = json.loads(param)
        excelFileFullName = os.path.join(BASE_FOLDER,parameters['book'])
        try:
            if (os.path.isfile(excelFileFullName)):
                wb = openpyxl.load_workbook(excelFileFullName,data_only=True)
                ws = wb[parameters['sheet']]
            else:
                return "File or sheet does not exist!"
            startCell = parameters['range'].split(':')[0]
            endCell = startCell if len(parameters['range'].split(':')) == 1 else parameters['range'].split(':')[1]
            cellRange = ws[startCell:endCell]
            result = []
            headers = parameters['header'].split(',')
            for row in cellRange:
                result.append({headers[i]:row[i].value for i in range(0,len(row))})
        except Exception as e:
            return e.message
        return json.dumps(result)
    def test_write_excel_function(self,param):
        '''Write data to excel from param like {"book":"write_excel.xlsx","sheet":"test","range":"A2:C3","data":[{"A2":"10","B2":"11","C2":"12"},{"A3":"30","B3":"31","C3":"32"}]}'''
        parameters = json.loads(param)
        excelFileFullName = os.path.join(BASE_FOLDER,parameters['book'])
        try:
            if (os.path.isfile(excelFileFullName)):
                return 'File exists!'
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
            ws.title = parameters['sheet']
            for row in parameters['data']:
               for k,v in row.items():
                  ws[k].value = v
            wb.save(excelFileFullName)
        except Exception as e:
            return e.message
        return 'success'

# UnitTest
if (platform.system() == 'Windows'):
    BASE_FOLDER = 'D:\Test'
if (platform.system() == 'Linux'):
    BASE_FOLDER = '/home/oracle'
#act = Action()
#act.test_excel_function()
#act.test_write_excel_function('{"book":"write_excel.xlsx","sheet":"test","range":"A2:C3","data":[{"A2":"10","B2":"11","C2":"12"},{"A3":"30","B3":"31","C3":"32"}]}')
#print(act.test_read_excel_function('{"book":"read_excel.xlsx","sheet":"test","range":"A2:C3","header":"a,b,c"}'))

class RequestHandler(BaseHTTPRequestHandler):
    server_version = "Excel Process Server/0.1"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def do_GET(self):
        message = ""
        if (self.path == '/status'):
            message = self.status()
        elif (self.path == '/test'):
            message = self.test()
        else:
            message = "Action not find."
        encoded = message.encode('utf-8', 'surrogateescape')
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-type", "text/html; charset=%s" % 'utf-8')
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)
        return

    def do_POST(self):

        length = self.headers.get('content-length')
        nbytes = int(length)
        
        if nbytes > MAX_BYTES:
            self.send_response(HTTPStatus.BAD_REQUEST)
            self.end_headers()
            self.wfile.write("File size over!")
            return

        rawPostData = self.rfile.read(nbytes)
        decodedPostData = rawPostData.decode('utf-8')
        #postData = urllib.parse.parse_qs(decodedPostData)
        postData = json.loads(decodedPostData)

        if postData is None or 'key' not in postData.keys() or postData['key'] != ACCESS_KEY:
            # 接続を強制切断とする
            return

        message = ""
        if (self.path == '/read'):
            message = self.read(postData)
        elif (self.path == '/write'):
            message = self.write(postData)
        else:
            message = "Action not find."

        encoded = message.encode('utf-8')
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-type", "text/plain; charset=%s" % 'utf-8')
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()

        self.wfile.write(encoded)
        return

    def status(self):
        return "Waiting..."
    def test(self):
        act = Action()
        return act.test_excel_function()
    def read(self,data):
        act = Action()
        message = ''
        try:
            fun = getattr(act,data['process_method'])
            message = json.dumps({'status':'success','data':fun(data['param'])})
        except:
            message = json.dumps({'status':'error','data':'Parameter process_method:{0} is not correct!'.format(data['process_method'])})
        return message
    def write(self,data):
        act = Action()
        message = ''
        try:
            fun = getattr(act,data['process_method'])
            message = json.dumps({'status':'success','data':fun(data['param'])})
        except:
            message = json.dumps({'status':'error','data':'Parameter process_method:{0} is not correct!'.format(data['process_method'])})
        return message
# -----------------------------------------------------------------------------
handler = RequestHandler
httpd = ThreadingHTTPServer(('',PORT),handler)
httpd.serve_forever()
# -----------------------------------------------------------------------------
# 参考情報
#   https://docs.python.org/ja/3/library/platform.html
#   https://docs.python.org/ja/3/library/http.server.html

# UnitTest From Powershell
# $postText = @{key="70949d2b88c04e869b074597700dd22d";fileName="test.xml";process_method="test_read_excel_function";param='{"book":"read_excel.xlsx","sheet":"test","range":"A2:C3","header":"a,b,c"}'} | ConvertTo-Json -Compress
# $postBody = [Text.Encoding]::UTF8.GetBytes($postText)
# Invoke-RestMethod -Method POST -Uri "http://localhost:10001/read" -Body $postBody -ContentType application/json
# $postText = @{key="70949d2b88c04e869b074597700dd22d";fileName="test.xml";process_method="test_write_excel_function";param='{"book":"write_excel.xlsx","sheet":"test","range":"A2:C3","header":"a,b,c","data":[{"A2":"10","B2":"11","C2":"12"},{"A3":"30","B3":"31","C3":"32"}]}'} | ConvertTo-Json -Compress
# $postBody = [Text.Encoding]::UTF8.GetBytes($postText)
# Invoke-RestMethod -Method POST -Uri "http://localhost:10001/write" -Body $postBody -ContentType application/json

# UnitTest From curl
# curl -H 'Content-Type:application/json; charset=utf-8' -d '{"key":"70949d2b88c04e869b074597700dd22d","fileName":"test.xml","process_method":"test_read_excel_function","param":"{\"book\":\"read_excel.xlsx\",\"sheet\":\"test\",\"range\":\"A2:C3\",\"header\":\"a,b,c\"}"}' http://localhost:10001/read
# curl -H 'Content-Type:application/json; charset=utf-8' -d '{"key":"70949d2b88c04e869b074597700dd22d","fileName":"test.xml","process_method":"test_write_excel_function","param":"{\"book\":\"write_excel.xlsx\",\"sheet\":\"test\",\"range\":\"A2:C3\",\"header\":\"a,b,c\",\"data\":[{\"A2\":\"10\",\"B2\":\"11\",\"C2\":\"12\"},{\"A3\":\"30\",\"B3\":\"31\",\"C3\":\"32\"}]}"}' http://localhost:10001/write